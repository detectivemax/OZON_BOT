import json

import openpyxl as openpyxl
from openpyxl.styles import Font, PatternFill

import requests as requests
from functions import get_num, get_date
from config.config import OZON_CLIENT_ID, OZON_API_KEY

from xlsxwriter.workbook import Workbook

# Цена ткани
def get_price(name):
    price_list = {
        "Лайн": 10.7,
        "Мальта": 12,
        "Бали": 21,
        "Рио": 22,
        "Сиде": 24,
        "Сфера": 28

    }

    if name in price_list:
        return price_list[name]

    else:
        return "-"


# Информация об отчетё
def get_ozon_response_info(code):
    # Метод
    link = "https://api-seller.ozon.ru/v1/report/info"

    head = {
        "Client-Id": OZON_CLIENT_ID,
        "Api-Key": OZON_API_KEY
    }

    # Параметры
    body = {
        "code": code
    }

    body = json.dumps(body)
    response = requests.post(link, headers=head, data=body)

    return response.json()["result"]


def get_ozon_list():
    link = "https://api-seller.ozon.ru/v1/report/list"

    info = requests.post(link,
                         headers={
                             "Client-Id": OZON_CLIENT_ID,
                             "Api-Key": OZON_API_KEY
                         },
                         )

    # print(info.json()["result"]["reports"])


def ozon_get_analytics_data(data_1, data_2):
    # Дата к виду озона
    data_1 = f"{data_1[6::]}-{data_1[3:5]}-{data_1[0:2]}"
    data_2 = f"{data_2[6::]}-{data_2[3:5]}-{data_2[0:2]}"

    # Метод
    link = "https://api-seller.ozon.ru/v1/analytics/data"

    head = {
        "Client-Id": OZON_CLIENT_ID,
        "Api-Key": OZON_API_KEY
    }

    # Параметры
    body = {
          "date_from": data_1,
          "date_to": data_2,

          "metrics": [
              "ordered_units",
              "revenue"
          ],

          "dimension": [

            "sku",
            "day"
          ],

          "limit": 1000,
          "offset": 0
    }

    body = json.dumps(body)
    response = requests.post(link, headers=head, data=body)

    return response.json()["result"]["data"]


# Указать даты продаж
def data_to_excel(ozon_content):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = 'Продажи за период'

    sheet["A2"] = "Дата"
    sheet.column_dimensions["A"].width = 20
    sheet["A2"].font = Font(bold=True)

    sheet["B2"] = "ID товара"
    sheet.column_dimensions["B"].width = 20
    sheet["B2"].font = Font(bold=True)

    sheet["C2"] = "Наименование товара"
    sheet.column_dimensions["C"].width = 80
    sheet["C2"].font = Font(bold=True)

    sheet["D2"] = "Продано позиций"
    sheet.column_dimensions["D"].width = 20
    sheet["D2"].font = Font(bold=True)

    sheet["E2"] = "Выручка"
    sheet.column_dimensions["E"].width = 15
    sheet["E2"].font = Font(bold=True)

    sheet["F2"] = "Комиссия за продажу"
    sheet.column_dimensions["F"].width = 25
    sheet["F2"].font = Font(bold=True)

    sheet["G2"] = "Обработка и достава"
    sheet.column_dimensions["G"].width = 25
    sheet["G2"].font = Font(bold=True)

    sheet["H2"] = "4%"
    sheet.column_dimensions["H"].width = 15
    sheet["H2"].font = Font(bold=True)

    sheet["I2"] = "Выручка после вычета комиссий"
    sheet.column_dimensions["I"].width = 35
    sheet["I2"].font = Font(bold=True)
    sheet["I2"].fill = PatternFill(start_color="99ff99", end_color="99ff99", fill_type="solid")

    sheet["J2"] = "Количество ламелей"
    sheet.column_dimensions["J"].width = 25
    sheet["J2"].font = Font(bold=True)

    sheet["K2"] = "Ширина"
    sheet.column_dimensions["K"].width = 15
    sheet["K2"].font = Font(bold=True)

    sheet["L2"] = "Высота"
    sheet.column_dimensions["L"].width = 15
    sheet["L2"].font = Font(bold=True)

    sheet["M2"] = "Ткань"
    sheet.column_dimensions["M"].width = 15
    sheet["M2"].font = Font(bold=True)

    sheet["N2"] = "Цена ткани"
    sheet.column_dimensions["N"].width = 15
    sheet["N2"].font = Font(bold=True)

    sheet["O2"] = "Себестоимость без работы"
    sheet.column_dimensions["O"].width = 30
    sheet["O2"].font = Font(bold=True)

    sheet["P2"] = "Себестоимость c работой"
    sheet.column_dimensions["P"].width = 30
    sheet["P2"].font = Font(bold=True)
    sheet["P2"].fill = PatternFill(start_color="99ff99", end_color="99ff99", fill_type="solid")

    sheet["Q2"] = "Прибыль общая"
    sheet.column_dimensions["Q"].width = 25
    sheet["Q2"].font = Font(bold=True)

    sheet["R2"] = "Максиму"
    sheet.column_dimensions["R"].width = 20
    sheet["R2"].font = Font(bold=True)

    sheet["S2"] = "Вернуть"
    sheet.column_dimensions["S"].width = 20
    sheet["S2"].font = Font(bold=True)
    sheet["S2"].fill = PatternFill(start_color="f8bf8f", end_color="f8bf8f", fill_type="solid")

    row_index = 3
    sum_total = 0
    for row in ozon_content:
        if row['metrics'][0] != 0:
            try:
                # Дата
                sheet[row_index][0].value = row['dimensions'][1]['id']

                # ID
                sheet[row_index][1].value = row['dimensions'][0]['id']

                # Наименование товара
                sheet[row_index][2].value = row['dimensions'][0]['name']

                # Продано позиций
                sheet[row_index][3].value = row['metrics'][0]

                # Выручка
                sheet[row_index][4].value = row['metrics'][1]

                # Комиссия за продажу
                com_1 = int(0.1*(int(row['metrics'][1])))
                sheet[row_index][5].value = com_1

                # Обработка и достава
                com_2 = int(0.115*(int(row['metrics'][1])))
                sheet[row_index][6].value = com_2

                # 4%
                com_3 = int(0.04*(int(row['metrics'][1])))
                sheet[row_index][7].value = com_3

                # Выручка после вычета комиссий
                cash = int(row['metrics'][1]) - com_1 - com_2 - com_3
                sheet[row_index][8].value = cash
                sheet[f"I{row_index}"].fill = PatternFill(start_color="99ff99", end_color="99ff99", fill_type="solid")

                ###
                pos = row['dimensions'][0]['name']
                if pos.find("Ламели для вертикальных жалюзи.") != -1:
                    pos = "Ламели для вертикальных жалюзи."

                if pos == "Ламели для вертикальных жалюзи.":
                    # Количество ламелей
                    name = row['dimensions'][0]['name']
                    lam = int(name[name.find(" х ") + 3: name.find(" шт.")])
                    sheet[row_index][9].value = lam
                    name = name.replace(f"{lam} шт", "")

                    # Ширина
                    width = lam * 8
                    sheet[row_index][10].value = width

                    # Высота
                    length = get_num(name)
                    sheet[row_index][11].value = length

                    # Ткань
                    name = name.replace("Ламели для вертикальных жалюзи. Ткань ", "")
                    name = name[: name.find(" ")]
                    sheet[row_index][12].value = name

                    # Цена ткани
                    price = get_price(name)
                    sheet[row_index][13].value = price

                    # Себестоимость без работы
                    price = 1.03 * ((lam * (length + 10) / 100 * price + 30) * row['metrics'][0])
                    sheet[row_index][14].value = price

                    # Себестоимость c работой
                    job = (width / 100 * length / 100) * 30 * row['metrics'][0]
                    price = price + job
                    sheet[row_index][15].value = price
                    sheet[f"P{row_index}"].fill = PatternFill(start_color="99ff99", end_color="99ff99",
                                                              fill_type="solid")

                    # Прибыть общая
                    value = cash - price
                    sheet[row_index][16].value = value

                    # Максиму
                    to_maxim = round(value / 3 + com_3)
                    sheet[row_index][17].value = to_maxim

                    # Вернуть
                    to_pay_back = cash - to_maxim
                    sheet[row_index][18].value = round(to_pay_back)
                    sheet[f"S{row_index}"].fill = PatternFill(start_color="f8bf8f", end_color="f8bf8f", fill_type="solid")




                row_index += 1
            except Exception as e:
                print(e)

            # print(row)
            # print(f"Название: {row['dimensions'][0]['name']}")
            # print(f"Дата: {row['dimensions'][1]['id']}")
            # print(f"ID: {row['dimensions'][0]['id']}")
            # print(f"Продано позиций: {row['metrics'][0]}")
            # print(f"Выручка: {row['metrics'][1]}")
            # print()

            sum_total += int(row['metrics'][1])

    # Продажи за месяц
    sheet.row_dimensions[1].height = 50
    sheet.row_dimensions[2].height = 50

    sheet["A1"] = "Продажи за период"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"] = str(sum_total)
    sheet["B1"].font = Font(bold=True)
    book.save("month_result.xlsx")
    book.close()
